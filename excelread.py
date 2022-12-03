from openpyxl import workbook
import openpyxl
from mimetypes import guess_type
from email.encoders import encode_base64
from win32com.client import Dispatch
import win32com

# current excel sheet location
path = "C:\\Users\\muniv\\Desktop\\python\\Book1.xlsx"
workbook = openpyxl.load_workbook(path, data_only=True)

sheet1 = workbook["Sheet1"]
sheet2 = workbook["Sheet2"]

maxr = sheet2.max_row
maxc = sheet2.max_column
#Copying data
for r in range (1, maxr + 1):
    for c in range (1, maxc +1):
        sheet1.cell(row=r,column=c).value = sheet2.cell(row=r,column=c).value

#creating new excelsheet
workbook.save("New_workbook.xlsx")        
print("Done")

# sneding an email 
def send_email():
    outlook = win32com.client.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = "xyz@xyz.com"
    mail.subject = 'Today Sales report'
    mail.Body = ' Hi Team, Please find the todays Sale report in attachments'
    attachment = "C:\\Users\\muniv\\Desktop\\python\\eod\\New_workbook.xlsx"
    mail.attachments.add(attachment)
    mail.send()

send_email()    